"""Excel export service using openpyxl"""
from datetime import date
from typing import List, Dict, Any
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Service for exporting reports to Excel files"""

    def __init__(self):
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill("solid", fgColor="1A5276")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.title_font = Font(bold=True, size=14)
        self.subtitle_font = Font(bold=True, size=11)

        self.currency_format = '#,##0.00'
        self.date_format = 'DD-MM-YYYY'

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_header_style(self, ws, row: int, start_col: int = 1, end_col: int = None):
        """Apply header styling to a row"""
        if end_col is None:
            end_col = ws.max_column

        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_sales_report(
        self,
        report_data: Dict,
        invoices: List,
        output_path: str,
        company_name: str = "GST Billing"
    ) -> Dict:
        """
        Export sales report to Excel

        Args:
            report_data: Dict with total_sales, total_tax, invoice_count, payment_breakdown
            invoices: List of Invoice objects
            output_path: Path to save Excel file
            company_name: Company name for header

        Returns:
            Dict with success status and details
        """
        try:
            wb = Workbook()

            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Title
            ws_summary['A1'] = f"{company_name} - Sales Report"
            ws_summary['A1'].font = self.title_font
            ws_summary.merge_cells('A1:D1')

            # Date range
            ws_summary['A2'] = f"Date: {report_data.get('date', date.today())}"
            ws_summary['A2'].font = self.subtitle_font

            # Summary data
            ws_summary['A4'] = "Metric"
            ws_summary['B4'] = "Value"
            self._apply_header_style(ws_summary, 4, 1, 2)

            summary_rows = [
                ("Total Sales", report_data.get('total_sales', 0)),
                ("Total Tax Collected", report_data.get('total_tax', 0)),
                ("Invoice Count", report_data.get('invoice_count', 0)),
            ]

            row = 5
            for metric, value in summary_rows:
                ws_summary.cell(row=row, column=1, value=metric).border = self.thin_border
                cell = ws_summary.cell(row=row, column=2, value=value)
                cell.border = self.thin_border
                if isinstance(value, (int, float)) and metric != "Invoice Count":
                    cell.number_format = self.currency_format
                row += 1

            # Payment breakdown
            row += 1
            ws_summary.cell(row=row, column=1, value="Payment Mode Breakdown").font = self.subtitle_font
            row += 1

            ws_summary.cell(row=row, column=1, value="Payment Mode")
            ws_summary.cell(row=row, column=2, value="Amount")
            self._apply_header_style(ws_summary, row, 1, 2)

            row += 1
            payment_breakdown = report_data.get('payment_breakdown', {})
            for mode, amount in payment_breakdown.items():
                ws_summary.cell(row=row, column=1, value=mode).border = self.thin_border
                cell = ws_summary.cell(row=row, column=2, value=amount)
                cell.border = self.thin_border
                cell.number_format = self.currency_format
                row += 1

            self._auto_adjust_columns(ws_summary)

            # Sheet 2: Invoice Details
            if invoices:
                ws_details = wb.create_sheet("Invoice Details")

                headers = ["Invoice No", "Date", "Customer", "Subtotal", "CGST", "SGST", "IGST", "Discount", "Grand Total", "Payment Mode"]

                for col, header in enumerate(headers, 1):
                    ws_details.cell(row=1, column=col, value=header)
                self._apply_header_style(ws_details, 1, 1, len(headers))

                row = 2
                for inv in invoices:
                    if inv.is_cancelled:
                        continue

                    ws_details.cell(row=row, column=1, value=inv.invoice_number).border = self.thin_border
                    ws_details.cell(row=row, column=2, value=str(inv.invoice_date)).border = self.thin_border
                    ws_details.cell(row=row, column=3, value=inv.customer_name or "Cash").border = self.thin_border

                    for col, value in enumerate([inv.subtotal, inv.cgst_total, inv.sgst_total, inv.igst_total, inv.discount, inv.grand_total], 4):
                        cell = ws_details.cell(row=row, column=col, value=value)
                        cell.border = self.thin_border
                        cell.number_format = self.currency_format

                    ws_details.cell(row=row, column=10, value=inv.payment_mode).border = self.thin_border
                    row += 1

                self._auto_adjust_columns(ws_details)

            wb.save(output_path)

            return {
                'success': True,
                'path': output_path,
                'sheets': ['Summary', 'Invoice Details'] if invoices else ['Summary']
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def export_gst_report(
        self,
        gst_summary: Dict,
        output_path: str,
        company_name: str = "GST Billing"
    ) -> Dict:
        """
        Export GST report to Excel

        Args:
            gst_summary: Dict with GST breakdown
            output_path: Path to save Excel file
            company_name: Company name for header

        Returns:
            Dict with success status and details
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "GST Report"

            # Title
            ws['A1'] = f"{company_name} - GST Report"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:E1')

            # Date range
            start_date = gst_summary.get('start_date', '')
            end_date = gst_summary.get('end_date', '')
            ws['A2'] = f"Period: {start_date} to {end_date}"
            ws['A2'].font = self.subtitle_font

            # Overall summary
            ws['A4'] = "Tax Summary"
            ws['A4'].font = self.subtitle_font

            ws['A5'] = "Description"
            ws['B5'] = "Amount"
            self._apply_header_style(ws, 5, 1, 2)

            summary_data = [
                ("Total Taxable Value", gst_summary.get('total_taxable', 0)),
                ("Total CGST", gst_summary.get('total_cgst', 0)),
                ("Total SGST", gst_summary.get('total_sgst', 0)),
                ("Total IGST", gst_summary.get('total_igst', 0)),
                ("Total Tax Collected", gst_summary.get('total_tax', 0)),
            ]

            row = 6
            for desc, amount in summary_data:
                ws.cell(row=row, column=1, value=desc).border = self.thin_border
                cell = ws.cell(row=row, column=2, value=amount)
                cell.border = self.thin_border
                cell.number_format = self.currency_format
                row += 1

            # Rate-wise breakdown
            row += 1
            ws.cell(row=row, column=1, value="Rate-wise Breakdown").font = self.subtitle_font
            row += 1

            headers = ["GST Rate (%)", "Taxable Value", "CGST", "SGST", "IGST", "Total Tax"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, 1, len(headers))

            row += 1
            rate_wise = gst_summary.get('rate_wise', {})
            for rate, data in sorted(rate_wise.items()):
                ws.cell(row=row, column=1, value=f"{rate}%").border = self.thin_border

                values = [
                    data.get('taxable', 0),
                    data.get('cgst', 0),
                    data.get('sgst', 0),
                    data.get('igst', 0),
                    data.get('cgst', 0) + data.get('sgst', 0) + data.get('igst', 0)
                ]

                for col, value in enumerate(values, 2):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.thin_border
                    cell.number_format = self.currency_format
                row += 1

            self._auto_adjust_columns(ws)
            wb.save(output_path)

            return {
                'success': True,
                'path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def export_stock_report(
        self,
        stock_items: List[Dict],
        output_path: str,
        company_name: str = "GST Billing"
    ) -> Dict:
        """
        Export stock report to Excel

        Args:
            stock_items: List of dicts with product stock info
            output_path: Path to save Excel file
            company_name: Company name for header

        Returns:
            Dict with success status and details
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Report"

            # Title
            ws['A1'] = f"{company_name} - Stock Report"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:F1')

            # Date
            ws['A2'] = f"Generated on: {date.today()}"
            ws['A2'].font = self.subtitle_font

            # Summary
            total_items = len(stock_items)
            low_stock_count = sum(1 for item in stock_items if item.get('is_low', False))
            total_value = sum(item.get('stock_value', 0) for item in stock_items)

            ws['A4'] = "Summary"
            ws['A4'].font = self.subtitle_font

            ws['A5'] = "Total Products"
            ws['B5'] = total_items
            ws['A6'] = "Low Stock Items"
            ws['B6'] = low_stock_count
            ws['A7'] = "Total Stock Value"
            cell = ws['B7']
            cell.value = total_value
            cell.number_format = self.currency_format

            # Stock details
            headers = ["Product Name", "HSN Code", "Unit", "Stock Qty", "Price", "Stock Value", "Status"]

            row = 9
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, 1, len(headers))

            row = 10
            low_fill = PatternFill("solid", fgColor="FFCCCC")  # Light red for low stock

            for item in stock_items:
                ws.cell(row=row, column=1, value=item.get('name', '')).border = self.thin_border
                ws.cell(row=row, column=2, value=item.get('hsn_code', '')).border = self.thin_border
                ws.cell(row=row, column=3, value=item.get('unit', '')).border = self.thin_border
                ws.cell(row=row, column=4, value=item.get('stock_qty', 0)).border = self.thin_border

                cell_price = ws.cell(row=row, column=5, value=item.get('price', 0))
                cell_price.border = self.thin_border
                cell_price.number_format = self.currency_format

                cell_value = ws.cell(row=row, column=6, value=item.get('stock_value', 0))
                cell_value.border = self.thin_border
                cell_value.number_format = self.currency_format

                status = "Low Stock" if item.get('is_low', False) else "OK"
                cell_status = ws.cell(row=row, column=7, value=status)
                cell_status.border = self.thin_border

                # Highlight low stock rows
                if item.get('is_low', False):
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = low_fill

                row += 1

            self._auto_adjust_columns(ws)
            wb.save(output_path)

            return {
                'success': True,
                'path': output_path,
                'total_items': total_items,
                'low_stock_count': low_stock_count
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def export_invoices_list(
        self,
        invoices: List,
        output_path: str,
        company_name: str = "GST Billing",
        title: str = "Invoice List"
    ) -> Dict:
        """
        Export list of invoices to Excel

        Args:
            invoices: List of Invoice objects
            output_path: Path to save Excel file
            company_name: Company name for header
            title: Report title

        Returns:
            Dict with success status and details
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"

            # Title
            ws['A1'] = f"{company_name} - {title}"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:J1')

            # Headers
            headers = [
                "Invoice No", "Date", "Customer", "GSTIN",
                "Subtotal", "CGST", "SGST", "IGST", "Discount", "Grand Total",
                "Payment Mode", "Status"
            ]

            row = 3
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, 1, len(headers))

            row = 4
            cancelled_fill = PatternFill("solid", fgColor="DDDDDD")

            for inv in invoices:
                ws.cell(row=row, column=1, value=inv.invoice_number).border = self.thin_border
                ws.cell(row=row, column=2, value=str(inv.invoice_date)).border = self.thin_border
                ws.cell(row=row, column=3, value=inv.customer_name or "Cash").border = self.thin_border
                ws.cell(row=row, column=4, value=getattr(inv, 'customer_gstin', '')).border = self.thin_border

                for col, value in enumerate([inv.subtotal, inv.cgst_total, inv.sgst_total, inv.igst_total, inv.discount, inv.grand_total], 5):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.thin_border
                    cell.number_format = self.currency_format

                ws.cell(row=row, column=11, value=inv.payment_mode).border = self.thin_border
                status = "Cancelled" if inv.is_cancelled else "Active"
                ws.cell(row=row, column=12, value=status).border = self.thin_border

                if inv.is_cancelled:
                    for col in range(1, 13):
                        ws.cell(row=row, column=col).fill = cancelled_fill

                row += 1

            self._auto_adjust_columns(ws)
            wb.save(output_path)

            return {
                'success': True,
                'path': output_path,
                'invoice_count': len(invoices)
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
