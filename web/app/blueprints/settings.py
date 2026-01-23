"""Settings blueprint - Company settings, email config, categories, backup/restore"""
from datetime import datetime, date
from io import BytesIO
import json
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from app.extensions import db
from app.models.company import Company
from app.models.category import Category
from app.models.product import Product
from app.models.customer import Customer
from app.models.invoice import Invoice, InvoiceItem
from functools import wraps

settings_bp = Blueprint('settings', __name__, url_prefix='/settings')


def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin():
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard.index'))
        return f(*args, **kwargs)
    return decorated_function


@settings_bp.route('/')
@login_required
@admin_required
def index():
    """Settings dashboard"""
    company = Company.get()
    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    return render_template(
        'settings/index.html',
        company=company,
        categories=categories
    )


@settings_bp.route('/company', methods=['GET', 'POST'])
@login_required
@admin_required
def company():
    """Company settings"""
    company = Company.get()

    if request.method == 'POST':
        if not company:
            company = Company()

        company.name = request.form.get('name', '').strip()
        company.address = request.form.get('address', '').strip()
        company.gstin = request.form.get('gstin', '').strip().upper()
        company.state_code = request.form.get('state_code', '32').strip()
        company.phone = request.form.get('phone', '').strip()
        company.email = request.form.get('email', '').strip()
        company.bank_name = request.form.get('bank_name', '').strip()
        company.bank_account = request.form.get('bank_account', '').strip()
        company.bank_ifsc = request.form.get('bank_ifsc', '').strip().upper()
        company.pan = request.form.get('pan', '').strip().upper()
        company.invoice_prefix = request.form.get('invoice_prefix', 'INV').strip().upper()
        company.invoice_terms = request.form.get('invoice_terms', '').strip()

        company.save()
        flash('Company settings saved successfully', 'success')
        return redirect(url_for('settings.company'))

    # State codes for dropdown
    state_codes = [
        ('01', '01 - Jammu & Kashmir'),
        ('02', '02 - Himachal Pradesh'),
        ('03', '03 - Punjab'),
        ('04', '04 - Chandigarh'),
        ('05', '05 - Uttarakhand'),
        ('06', '06 - Haryana'),
        ('07', '07 - Delhi'),
        ('08', '08 - Rajasthan'),
        ('09', '09 - Uttar Pradesh'),
        ('10', '10 - Bihar'),
        ('11', '11 - Sikkim'),
        ('12', '12 - Arunachal Pradesh'),
        ('13', '13 - Nagaland'),
        ('14', '14 - Manipur'),
        ('15', '15 - Mizoram'),
        ('16', '16 - Tripura'),
        ('17', '17 - Meghalaya'),
        ('18', '18 - Assam'),
        ('19', '19 - West Bengal'),
        ('20', '20 - Jharkhand'),
        ('21', '21 - Odisha'),
        ('22', '22 - Chhattisgarh'),
        ('23', '23 - Madhya Pradesh'),
        ('24', '24 - Gujarat'),
        ('26', '26 - Dadra & Nagar Haveli'),
        ('27', '27 - Maharashtra'),
        ('28', '28 - Andhra Pradesh'),
        ('29', '29 - Karnataka'),
        ('30', '30 - Goa'),
        ('31', '31 - Lakshadweep'),
        ('32', '32 - Kerala'),
        ('33', '33 - Tamil Nadu'),
        ('34', '34 - Puducherry'),
        ('35', '35 - Andaman & Nicobar'),
        ('36', '36 - Telangana'),
        ('37', '37 - Andhra Pradesh (New)'),
        ('38', '38 - Ladakh'),
    ]

    return render_template(
        'settings/company.html',
        company=company,
        state_codes=state_codes
    )


@settings_bp.route('/email', methods=['GET', 'POST'])
@login_required
@admin_required
def email():
    """Email settings"""
    company = Company.get()

    if request.method == 'POST':
        if not company:
            company = Company()

        company.smtp_server = request.form.get('smtp_server', '').strip()
        company.smtp_port = int(request.form.get('smtp_port', 587) or 587)
        company.smtp_username = request.form.get('smtp_username', '').strip()
        smtp_password = request.form.get('smtp_password', '').strip()
        if smtp_password:  # Only update if provided
            company.smtp_password = smtp_password
        company.smtp_use_tls = request.form.get('smtp_use_tls') == '1'
        company.email_from = request.form.get('email_from', '').strip()
        company.admin_notification_email = request.form.get('admin_notification_email', '').strip()

        company.save()
        flash('Email settings saved successfully', 'success')
        return redirect(url_for('settings.email'))

    return render_template('settings/email.html', company=company)


@settings_bp.route('/email/test', methods=['GET', 'POST'])
@login_required
def test_email():
    """Test email configuration"""
    try:
        # Handle GET requests - redirect to email settings
        if request.method == 'GET':
            flash('Please use the test email form on this page', 'info')
            return redirect(url_for('settings.email'))

        # Check admin
        if hasattr(current_user, 'is_admin') and callable(current_user.is_admin):
            if not current_user.is_admin():
                flash('Admin access required', 'error')
                return redirect(url_for('settings.email'))

        company = Company.get()
        if not company or not company.smtp_server:
            flash('Please configure email settings first', 'error')
            return redirect(url_for('settings.email'))

        test_to = request.form.get('test_email', '').strip()
        if not test_to:
            flash('Please enter a test email address', 'error')
            return redirect(url_for('settings.email'))

        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        msg = MIMEMultipart()
        msg['From'] = company.email_from or company.smtp_username
        msg['To'] = test_to
        msg['Subject'] = 'Cosmic Surgical - Test Email'

        body = f"""
        This is a test email from Cosmic Surgical Billing.

        If you received this email, your email configuration is working correctly.

        Company: {company.name}
        SMTP Server: {company.smtp_server}
        """
        msg.attach(MIMEText(body, 'plain'))

        # Connect to SMTP server with proper handshake
        if company.smtp_use_tls:
            server = smtplib.SMTP(company.smtp_server, company.smtp_port, timeout=30)
            server.ehlo()  # Required for Gmail
            server.starttls()
            server.ehlo()  # Required after starttls
        else:
            server = smtplib.SMTP_SSL(company.smtp_server, company.smtp_port, timeout=30)

        server.login(company.smtp_username, company.smtp_password)
        server.send_message(msg)
        server.quit()

        flash(f'Test email sent successfully to {test_to}', 'success')
    except smtplib.SMTPAuthenticationError as e:
        flash(f'Authentication failed: Check username and app password. Error: {str(e)}', 'error')
    except smtplib.SMTPConnectError as e:
        flash(f'Could not connect to SMTP server: {str(e)}', 'error')
    except smtplib.SMTPException as e:
        flash(f'SMTP Error: {str(e)}', 'error')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('settings.email'))


# Category Management
@settings_bp.route('/categories')
@login_required
@admin_required
def categories():
    """Category management"""
    categories = Category.query.order_by(Category.name).all()
    return render_template('settings/categories.html', categories=categories)


@settings_bp.route('/categories/add', methods=['POST'])
@login_required
@admin_required
def add_category():
    """Add new category"""
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()

    if not name:
        flash('Category name is required', 'error')
        return redirect(url_for('settings.categories'))

    if Category.query.filter_by(name=name).first():
        flash('Category already exists', 'error')
        return redirect(url_for('settings.categories'))

    category = Category(name=name, description=description)
    db.session.add(category)
    db.session.commit()

    flash(f'Category "{name}" created successfully', 'success')
    return redirect(url_for('settings.categories'))


@settings_bp.route('/categories/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def edit_category(id):
    """Edit category"""
    category = Category.query.get_or_404(id)

    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()

    if not name:
        flash('Category name is required', 'error')
        return redirect(url_for('settings.categories'))

    existing = Category.query.filter_by(name=name).first()
    if existing and existing.id != id:
        flash('Category name already in use', 'error')
        return redirect(url_for('settings.categories'))

    category.name = name
    category.description = description
    db.session.commit()

    flash(f'Category "{name}" updated successfully', 'success')
    return redirect(url_for('settings.categories'))


@settings_bp.route('/categories/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_category(id):
    """Toggle category active status"""
    category = Category.query.get_or_404(id)
    category.is_active = not category.is_active
    db.session.commit()

    status = 'activated' if category.is_active else 'deactivated'
    flash(f'Category "{category.name}" {status}', 'success')
    return redirect(url_for('settings.categories'))


# ==================== Bulk Import/Export ====================

@settings_bp.route('/import-export')
@login_required
@admin_required
def import_export():
    """Import/Export data page"""
    return render_template('settings/import_export.html')


@settings_bp.route('/export/products')
@login_required
@admin_required
def export_products():
    """Export all products to Excel"""
    products = Product.query.order_by(Product.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Headers
    headers = ['Name', 'Barcode', 'HSN Code', 'Category', 'Price', 'Cost Price',
               'GST Rate', 'Stock Qty', 'Unit', 'Low Stock Alert', 'Active']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for p in products:
        ws.append([
            p.name,
            p.barcode or '',
            p.hsn_code or '',
            p.category.name if p.category else '',
            p.price,
            p.cost_price if hasattr(p, 'cost_price') else 0,
            p.gst_rate,
            p.stock_qty,
            p.unit,
            p.low_stock_alert,
            'Yes' if p.is_active else 'No'
        ])

    # Auto-width columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'products_export_{date.today().strftime("%Y%m%d")}.xlsx'
    )


@settings_bp.route('/export/customers')
@login_required
@admin_required
def export_customers():
    """Export all customers to Excel"""
    customers = Customer.query.order_by(Customer.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = ['Name', 'Phone', 'Email', 'Address', 'GSTIN', 'State Code', 'State Name', 'Active']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for c in customers:
        ws.append([
            c.name,
            c.phone or '',
            c.email or '',
            c.address or '',
            c.gstin or '',
            c.state_code or '',
            c.state_name or '',
            'Yes' if c.is_active else 'No'
        ])

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'customers_export_{date.today().strftime("%Y%m%d")}.xlsx'
    )


@settings_bp.route('/import/products', methods=['POST'])
@login_required
@admin_required
def import_products():
    """Import products from Excel"""
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('settings.import_export'))

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('settings.import_export'))

    try:
        wb = load_workbook(file)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]

        imported = 0
        updated = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))

                name = str(data.get('name', '')).strip()
                if not name:
                    continue

                barcode = str(data.get('barcode', '')).strip()

                # Check if product exists (by barcode or name)
                product = None
                if barcode:
                    product = Product.query.filter_by(barcode=barcode).first()
                if not product:
                    product = Product.query.filter_by(name=name).first()

                if product:
                    # Update existing
                    product.name = name
                    product.barcode = barcode
                    product.hsn_code = str(data.get('hsn code', data.get('hsn_code', ''))).strip()
                    product.price = float(data.get('price', 0) or 0)
                    product.gst_rate = float(data.get('gst rate', data.get('gst_rate', 18)) or 18)
                    product.stock_qty = float(data.get('stock qty', data.get('stock_qty', 0)) or 0)
                    product.unit = str(data.get('unit', 'NOS')).strip() or 'NOS'
                    product.low_stock_alert = int(data.get('low stock alert', data.get('low_stock_alert', 10)) or 10)
                    updated += 1
                else:
                    # Create new
                    product = Product(
                        name=name,
                        barcode=barcode,
                        hsn_code=str(data.get('hsn code', data.get('hsn_code', ''))).strip(),
                        price=float(data.get('price', 0) or 0),
                        gst_rate=float(data.get('gst rate', data.get('gst_rate', 18)) or 18),
                        stock_qty=float(data.get('stock qty', data.get('stock_qty', 0)) or 0),
                        unit=str(data.get('unit', 'NOS')).strip() or 'NOS',
                        low_stock_alert=int(data.get('low stock alert', data.get('low_stock_alert', 10)) or 10)
                    )
                    db.session.add(product)
                    imported += 1

            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        db.session.commit()

        msg = f'Import complete: {imported} new products, {updated} updated'
        if errors:
            msg += f', {len(errors)} errors'
        flash(msg, 'success' if not errors else 'warning')

    except Exception as e:
        flash(f'Error reading file: {str(e)}', 'error')

    return redirect(url_for('settings.import_export'))


@settings_bp.route('/import/customers', methods=['POST'])
@login_required
@admin_required
def import_customers():
    """Import customers from Excel"""
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('settings.import_export'))

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('settings.import_export'))

    try:
        wb = load_workbook(file)
        ws = wb.active

        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]

        imported = 0
        updated = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))

                name = str(data.get('name', '')).strip()
                if not name:
                    continue

                phone = str(data.get('phone', '')).strip()

                # Check if customer exists
                customer = None
                if phone:
                    customer = Customer.query.filter_by(phone=phone).first()
                if not customer:
                    customer = Customer.query.filter_by(name=name).first()

                if customer:
                    customer.name = name
                    customer.phone = phone
                    customer.email = str(data.get('email', '')).strip()
                    customer.address = str(data.get('address', '')).strip()
                    customer.gstin = str(data.get('gstin', '')).strip().upper()
                    customer.state_code = str(data.get('state code', data.get('state_code', ''))).strip()
                    customer.state_name = str(data.get('state name', data.get('state_name', ''))).strip()
                    updated += 1
                else:
                    customer = Customer(
                        name=name,
                        phone=phone,
                        email=str(data.get('email', '')).strip(),
                        address=str(data.get('address', '')).strip(),
                        gstin=str(data.get('gstin', '')).strip().upper(),
                        state_code=str(data.get('state code', data.get('state_code', ''))).strip(),
                        state_name=str(data.get('state name', data.get('state_name', ''))).strip()
                    )
                    db.session.add(customer)
                    imported += 1

            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        db.session.commit()

        msg = f'Import complete: {imported} new customers, {updated} updated'
        if errors:
            msg += f', {len(errors)} errors'
        flash(msg, 'success' if not errors else 'warning')

    except Exception as e:
        flash(f'Error reading file: {str(e)}', 'error')

    return redirect(url_for('settings.import_export'))


@settings_bp.route('/download/template/<template_type>')
@login_required
@admin_required
def download_template(template_type):
    """Download import template"""
    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    if template_type == 'products':
        ws.title = "Products"
        headers = ['Name', 'Barcode', 'HSN Code', 'Price', 'GST Rate', 'Stock Qty', 'Unit', 'Low Stock Alert']
        ws.append(headers)
        # Example row
        ws.append(['Sample Product', '1234567890', '1234', 100, 18, 50, 'NOS', 10])
        filename = 'products_import_template.xlsx'

    elif template_type == 'customers':
        ws.title = "Customers"
        headers = ['Name', 'Phone', 'Email', 'Address', 'GSTIN', 'State Code', 'State Name']
        ws.append(headers)
        ws.append(['Sample Customer', '9876543210', 'customer@example.com', '123 Main St', '32ABCDE1234F1Z5', '32', 'Kerala'])
        filename = 'customers_import_template.xlsx'

    else:
        flash('Invalid template type', 'error')
        return redirect(url_for('settings.import_export'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== Backup/Restore ====================

@settings_bp.route('/backup')
@login_required
@admin_required
def backup_page():
    """Backup/Restore page"""
    return render_template('settings/backup.html')


@settings_bp.route('/backup/create')
@login_required
@admin_required
def create_backup():
    """Create database backup as JSON"""
    backup_data = {
        'backup_date': datetime.utcnow().isoformat(),
        'version': '1.0',
        'company': None,
        'categories': [],
        'products': [],
        'customers': [],
        'invoices': []
    }

    # Company
    company = Company.get()
    if company:
        backup_data['company'] = company.to_dict()

    # Categories
    for cat in Category.query.all():
        backup_data['categories'].append({
            'id': cat.id,
            'name': cat.name,
            'description': cat.description,
            'is_active': cat.is_active
        })

    # Products
    for p in Product.query.all():
        backup_data['products'].append(p.to_dict())

    # Customers
    for c in Customer.query.all():
        backup_data['customers'].append(c.to_dict())

    # Invoices with items
    for inv in Invoice.query.all():
        inv_data = inv.to_dict()
        inv_data['items'] = [item.to_dict() for item in inv.items]
        backup_data['invoices'].append(inv_data)

    # Create JSON file
    output = BytesIO()
    output.write(json.dumps(backup_data, indent=2, default=str).encode('utf-8'))
    output.seek(0)

    filename = f'gst_billing_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

    return send_file(
        output,
        mimetype='application/json',
        as_attachment=True,
        download_name=filename
    )


@settings_bp.route('/backup/restore', methods=['POST'])
@login_required
@admin_required
def restore_backup():
    """Restore from backup JSON file"""
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('settings.backup_page'))

    file = request.files['file']
    if not file.filename.endswith('.json'):
        flash('Please upload a JSON backup file', 'error')
        return redirect(url_for('settings.backup_page'))

    try:
        backup_data = json.load(file)

        if 'version' not in backup_data:
            flash('Invalid backup file format', 'error')
            return redirect(url_for('settings.backup_page'))

        # Restore Categories
        if backup_data.get('categories'):
            for cat_data in backup_data['categories']:
                cat = Category.query.filter_by(name=cat_data['name']).first()
                if not cat:
                    cat = Category(name=cat_data['name'])
                cat.description = cat_data.get('description', '')
                cat.is_active = cat_data.get('is_active', True)
                db.session.add(cat)

        # Restore Products
        if backup_data.get('products'):
            for p_data in backup_data['products']:
                product = None
                if p_data.get('barcode'):
                    product = Product.query.filter_by(barcode=p_data['barcode']).first()
                if not product:
                    product = Product.query.filter_by(name=p_data['name']).first()
                if not product:
                    product = Product(name=p_data['name'])

                product.barcode = p_data.get('barcode', '')
                product.hsn_code = p_data.get('hsn_code', '')
                product.price = p_data.get('price', 0)
                product.gst_rate = p_data.get('gst_rate', 18)
                product.stock_qty = p_data.get('stock_qty', 0)
                product.unit = p_data.get('unit', 'NOS')
                product.low_stock_alert = p_data.get('low_stock_alert', 10)
                db.session.add(product)

        # Restore Customers
        if backup_data.get('customers'):
            for c_data in backup_data['customers']:
                customer = None
                if c_data.get('phone'):
                    customer = Customer.query.filter_by(phone=c_data['phone']).first()
                if not customer and c_data.get('gstin'):
                    customer = Customer.query.filter_by(gstin=c_data['gstin']).first()
                if not customer:
                    customer = Customer.query.filter_by(name=c_data['name']).first()
                if not customer:
                    customer = Customer(name=c_data['name'])

                customer.phone = c_data.get('phone', '')
                customer.email = c_data.get('email', '')
                customer.address = c_data.get('address', '')
                customer.gstin = c_data.get('gstin', '')
                customer.state_code = c_data.get('state_code', '')
                customer.state_name = c_data.get('state_name', '')
                db.session.add(customer)

        db.session.commit()

        flash('Backup restored successfully! Categories, Products, and Customers have been imported.', 'success')

    except json.JSONDecodeError:
        flash('Invalid JSON file', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Error restoring backup: {str(e)}', 'error')

    return redirect(url_for('settings.backup_page'))
