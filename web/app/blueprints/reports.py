"""Reports blueprint - Sales, GST, Stock, and GSTR-1 reports"""
from datetime import date, datetime, timedelta
from io import BytesIO
from flask import Blueprint, render_template, request, send_file, flash, jsonify
from flask_login import login_required
from sqlalchemy import func, extract
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.extensions import db
from app.models.invoice import Invoice, InvoiceItem
from app.models.credit_note import CreditNote, CreditNoteItem
from app.models.product import Product
from app.models.customer import Customer
from app.models.company import Company

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


def get_financial_year_dates(fy_year=None):
    """Get start and end dates of financial year (April to March)"""
    today = date.today()
    if fy_year is None:
        if today.month >= 4:
            fy_year = today.year
        else:
            fy_year = today.year - 1

    start_date = date(fy_year, 4, 1)
    end_date = date(fy_year + 1, 3, 31)
    return start_date, end_date, f"{fy_year}-{str(fy_year + 1)[-2:]}"


@reports_bp.route('/')
@login_required
def index():
    """Reports dashboard"""
    return render_template('reports/index.html')


@reports_bp.route('/sales')
@login_required
def sales_report():
    """Sales report - daily, monthly, yearly summaries"""
    report_type = request.args.get('type', 'daily')
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    # Default date range
    today = date.today()
    if not start_date_str:
        if report_type == 'daily':
            start_date = today - timedelta(days=30)
        elif report_type == 'monthly':
            start_date = date(today.year, 1, 1)
        else:  # yearly
            start_date = date(today.year - 2, 4, 1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Get invoices in range
    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date.desc()).all()

    # Calculate summary
    total_sales = sum(inv.grand_total for inv in invoices)
    total_tax = sum(inv.cgst_total + inv.sgst_total + inv.igst_total for inv in invoices)
    invoice_count = len(invoices)

    # Group by date/month/year
    sales_data = []
    if report_type == 'daily':
        grouped = {}
        for inv in invoices:
            key = inv.invoice_date
            if key not in grouped:
                grouped[key] = {'date': key, 'count': 0, 'subtotal': 0, 'tax': 0, 'total': 0}
            grouped[key]['count'] += 1
            grouped[key]['subtotal'] += inv.subtotal
            grouped[key]['tax'] += inv.cgst_total + inv.sgst_total + inv.igst_total
            grouped[key]['total'] += inv.grand_total
        sales_data = sorted(grouped.values(), key=lambda x: x['date'], reverse=True)

    elif report_type == 'monthly':
        grouped = {}
        for inv in invoices:
            key = date(inv.invoice_date.year, inv.invoice_date.month, 1)
            if key not in grouped:
                grouped[key] = {'date': key, 'count': 0, 'subtotal': 0, 'tax': 0, 'total': 0}
            grouped[key]['count'] += 1
            grouped[key]['subtotal'] += inv.subtotal
            grouped[key]['tax'] += inv.cgst_total + inv.sgst_total + inv.igst_total
            grouped[key]['total'] += inv.grand_total
        sales_data = sorted(grouped.values(), key=lambda x: x['date'], reverse=True)

    else:  # yearly
        grouped = {}
        for inv in invoices:
            # Financial year grouping
            if inv.invoice_date.month >= 4:
                fy = inv.invoice_date.year
            else:
                fy = inv.invoice_date.year - 1
            key = f"{fy}-{str(fy+1)[-2:]}"
            if key not in grouped:
                grouped[key] = {'year': key, 'count': 0, 'subtotal': 0, 'tax': 0, 'total': 0}
            grouped[key]['count'] += 1
            grouped[key]['subtotal'] += inv.subtotal
            grouped[key]['tax'] += inv.cgst_total + inv.sgst_total + inv.igst_total
            grouped[key]['total'] += inv.grand_total
        sales_data = sorted(grouped.values(), key=lambda x: x['year'], reverse=True)

    return render_template(
        'reports/sales.html',
        report_type=report_type,
        start_date=start_date,
        end_date=end_date,
        sales_data=sales_data,
        total_sales=total_sales,
        total_tax=total_tax,
        invoice_count=invoice_count
    )


@reports_bp.route('/gst')
@login_required
def gst_report():
    """GST report - tax summary by rate"""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    # Default to current month
    today = date.today()
    if not start_date_str:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Get all invoice items in range
    items = db.session.query(InvoiceItem).join(Invoice).filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).all()

    # Group by GST rate
    tax_by_rate = {}
    for item in items:
        rate = item.gst_rate
        if rate not in tax_by_rate:
            tax_by_rate[rate] = {
                'rate': rate,
                'taxable_value': 0,
                'cgst': 0,
                'sgst': 0,
                'igst': 0,
                'total_tax': 0,
                'count': 0
            }
        tax_by_rate[rate]['taxable_value'] += item.taxable_value
        tax_by_rate[rate]['cgst'] += item.cgst
        tax_by_rate[rate]['sgst'] += item.sgst
        tax_by_rate[rate]['igst'] += item.igst
        tax_by_rate[rate]['total_tax'] += item.cgst + item.sgst + item.igst
        tax_by_rate[rate]['count'] += 1

    gst_summary = sorted(tax_by_rate.values(), key=lambda x: x['rate'])

    # Totals
    total_taxable = sum(g['taxable_value'] for g in gst_summary)
    total_cgst = sum(g['cgst'] for g in gst_summary)
    total_sgst = sum(g['sgst'] for g in gst_summary)
    total_igst = sum(g['igst'] for g in gst_summary)
    total_tax = total_cgst + total_sgst + total_igst

    return render_template(
        'reports/gst.html',
        start_date=start_date,
        end_date=end_date,
        gst_summary=gst_summary,
        total_taxable=total_taxable,
        total_cgst=total_cgst,
        total_sgst=total_sgst,
        total_igst=total_igst,
        total_tax=total_tax
    )


@reports_bp.route('/stock')
@login_required
def stock_report():
    """Stock report - current inventory levels"""
    search = request.args.get('search', '').strip()
    low_stock = request.args.get('low_stock', '') == '1'
    category_id = request.args.get('category', type=int)

    query = Product.query.filter_by(is_active=True)

    if search:
        search_term = f'%{search}%'
        query = query.filter(
            db.or_(
                Product.name.ilike(search_term),
                Product.barcode.ilike(search_term),
                Product.hsn_code.ilike(search_term)
            )
        )

    if low_stock:
        query = query.filter(Product.stock_qty <= Product.low_stock_alert)

    if category_id:
        query = query.filter_by(category_id=category_id)

    products = query.order_by(Product.name).all()

    # Calculate totals
    total_stock_value = sum(p.stock_qty * p.price for p in products)
    total_items = len(products)
    low_stock_count = sum(1 for p in products if p.stock_qty <= p.low_stock_alert)

    # Get categories for filter
    from app.models.category import Category
    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()

    return render_template(
        'reports/stock.html',
        products=products,
        search=search,
        low_stock=low_stock,
        category_id=category_id,
        categories=categories,
        total_stock_value=total_stock_value,
        total_items=total_items,
        low_stock_count=low_stock_count
    )


@reports_bp.route('/gstr1')
@login_required
def gstr1_report():
    """GSTR-1 report - monthly/quarterly GST return data"""
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    today = date.today()
    if not month:
        month = today.month
    if not year:
        year = today.year

    # Get first and last day of month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # Get all invoices for the month
    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date, Invoice.invoice_number).all()

    # B2B (Business to Business) - invoices to registered dealers
    b2b_invoices = [inv for inv in invoices if inv.customer and inv.customer.gstin]

    # B2C Large (> 2.5 lakh inter-state to unregistered)
    # B2C Small (all other B2C)
    b2c_invoices = [inv for inv in invoices if not (inv.customer and inv.customer.gstin)]

    # Credit notes
    credit_notes = CreditNote.query.filter(
        CreditNote.credit_note_date.between(start_date, end_date),
        CreditNote.status != 'CANCELLED'
    ).order_by(CreditNote.credit_note_date).all()

    # Summary
    b2b_total = sum(inv.grand_total for inv in b2b_invoices)
    b2c_total = sum(inv.grand_total for inv in b2c_invoices)
    cn_total = sum(cn.grand_total for cn in credit_notes)

    total_taxable = sum(inv.subtotal for inv in invoices)
    total_cgst = sum(inv.cgst_total for inv in invoices)
    total_sgst = sum(inv.sgst_total for inv in invoices)
    total_igst = sum(inv.igst_total for inv in invoices)

    return render_template(
        'reports/gstr1.html',
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
        b2b_invoices=b2b_invoices,
        b2c_invoices=b2c_invoices,
        credit_notes=credit_notes,
        b2b_total=b2b_total,
        b2c_total=b2c_total,
        cn_total=cn_total,
        total_taxable=total_taxable,
        total_cgst=total_cgst,
        total_sgst=total_sgst,
        total_igst=total_igst
    )


@reports_bp.route('/gstr1/export')
@login_required
def export_gstr1():
    """Export GSTR-1 data to Excel"""
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    today = date.today()
    if not month:
        month = today.month
    if not year:
        year = today.year

    # Get first and last day of month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    company = Company.get()

    # Get invoices
    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date, Invoice.invoice_number).all()

    credit_notes = CreditNote.query.filter(
        CreditNote.credit_note_date.between(start_date, end_date),
        CreditNote.status != 'CANCELLED'
    ).order_by(CreditNote.credit_note_date).all()

    # Create workbook
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # B2B Sheet
    ws_b2b = wb.active
    ws_b2b.title = "B2B"
    b2b_headers = [
        "GSTIN/UIN of Recipient", "Invoice Number", "Invoice Date", "Invoice Value",
        "Place of Supply", "Reverse Charge", "Invoice Type", "E-Commerce GSTIN",
        "Rate", "Taxable Value", "Cess Amount"
    ]
    ws_b2b.append(b2b_headers)
    for cell in ws_b2b[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for inv in invoices:
        if inv.customer and inv.customer.gstin:
            for item in inv.items:
                ws_b2b.append([
                    inv.customer.gstin,
                    inv.invoice_number,
                    inv.invoice_date.strftime('%d-%m-%Y'),
                    inv.grand_total,
                    inv.customer.state_code if inv.customer else '',
                    'N',
                    'Regular',
                    '',
                    item.gst_rate,
                    item.taxable_value,
                    0
                ])

    # B2C Sheet
    ws_b2c = wb.create_sheet("B2C")
    b2c_headers = [
        "Type", "Place of Supply", "Rate", "Taxable Value",
        "Cess Amount", "E-Commerce GSTIN"
    ]
    ws_b2c.append(b2c_headers)
    for cell in ws_b2c[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Aggregate B2C by state and rate
    b2c_summary = {}
    for inv in invoices:
        if not (inv.customer and inv.customer.gstin):
            state = inv.customer.state_code if inv.customer else (company.state_code if company else '32')
            for item in inv.items:
                key = (state, item.gst_rate)
                if key not in b2c_summary:
                    b2c_summary[key] = 0
                b2c_summary[key] += item.taxable_value

    for (state, rate), taxable in b2c_summary.items():
        ws_b2c.append([
            'OE',  # OE = Others
            state,
            rate,
            taxable,
            0,
            ''
        ])

    # Credit/Debit Notes Sheet
    ws_cdn = wb.create_sheet("CDN")
    cdn_headers = [
        "GSTIN/UIN of Recipient", "Note Number", "Note Date", "Note Type",
        "Place of Supply", "Original Invoice Number", "Original Invoice Date",
        "Note Value", "Rate", "Taxable Value", "Cess Amount"
    ]
    ws_cdn.append(cdn_headers)
    for cell in ws_cdn[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for cn in credit_notes:
        customer_gstin = cn.customer.gstin if cn.customer else ''
        state_code = cn.customer.state_code if cn.customer else ''
        orig_inv = cn.original_invoice
        orig_date = orig_inv.invoice_date.strftime('%d-%m-%Y') if orig_inv else ''

        for item in cn.items:
            ws_cdn.append([
                customer_gstin,
                cn.credit_note_number,
                cn.credit_note_date.strftime('%d-%m-%Y'),
                'C',  # C = Credit Note
                state_code,
                cn.original_invoice_number,
                orig_date,
                cn.grand_total,
                item.gst_rate,
                item.taxable_value,
                0
            ])

    # HSN Summary Sheet
    ws_hsn = wb.create_sheet("HSN")
    hsn_headers = [
        "HSN", "Description", "UQC", "Total Quantity",
        "Total Value", "Taxable Value", "Integrated Tax",
        "Central Tax", "State Tax", "Cess"
    ]
    ws_hsn.append(hsn_headers)
    for cell in ws_hsn[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Aggregate by HSN
    hsn_summary = {}
    for inv in invoices:
        for item in inv.items:
            hsn = item.hsn_code or 'NA'
            if hsn not in hsn_summary:
                hsn_summary[hsn] = {
                    'description': item.product_name,
                    'unit': item.unit,
                    'qty': 0,
                    'total_value': 0,
                    'taxable': 0,
                    'igst': 0,
                    'cgst': 0,
                    'sgst': 0
                }
            hsn_summary[hsn]['qty'] += item.qty
            hsn_summary[hsn]['total_value'] += item.total
            hsn_summary[hsn]['taxable'] += item.taxable_value
            hsn_summary[hsn]['igst'] += item.igst
            hsn_summary[hsn]['cgst'] += item.cgst
            hsn_summary[hsn]['sgst'] += item.sgst

    for hsn, data in hsn_summary.items():
        ws_hsn.append([
            hsn,
            data['description'][:50],
            data['unit'],
            data['qty'],
            data['total_value'],
            data['taxable'],
            data['igst'],
            data['cgst'],
            data['sgst'],
            0
        ])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month_name = start_date.strftime('%B')
    filename = f"GSTR1_{month_name}_{year}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== GSTR-3B Report ====================

@reports_bp.route('/gstr3b')
@login_required
def gstr3b_report():
    """GSTR-3B summary report - Monthly return filing summary"""
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)

    # Calculate date range for the month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # Get all non-cancelled invoices for the month
    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).all()

    # Get credit notes for the month
    credit_notes = CreditNote.query.filter(
        CreditNote.credit_note_date.between(start_date, end_date),
        CreditNote.status != 'CANCELLED'
    ).all()

    # 3.1 Outward Supplies - Tax liability
    # (a) Outward taxable supplies (other than zero rated, nil rated and exempted)
    taxable_outward = {
        'taxable_value': 0,
        'igst': 0,
        'cgst': 0,
        'sgst': 0,
        'cess': 0
    }

    # (b) Outward taxable supplies (zero rated)
    zero_rated = {'taxable_value': 0, 'igst': 0}

    # (c) Other outward supplies (nil rated, exempted)
    exempt_nil = {'taxable_value': 0}

    # (d) Inward supplies (liable to reverse charge)
    reverse_charge = {
        'taxable_value': 0,
        'igst': 0,
        'cgst': 0,
        'sgst': 0,
        'cess': 0
    }

    # Process invoices
    for inv in invoices:
        is_rcm = getattr(inv, 'is_reverse_charge', False)

        if is_rcm:
            # Reverse charge - goes to 3.1(d)
            reverse_charge['taxable_value'] += inv.subtotal or 0
            reverse_charge['igst'] += inv.igst_total or 0
            reverse_charge['cgst'] += inv.cgst_total or 0
            reverse_charge['sgst'] += inv.sgst_total or 0
        else:
            # Check if all items are 0% GST (exempt/nil rated)
            items = list(inv.items) if hasattr(inv, 'items') else []
            all_exempt = all(item.gst_rate == 0 for item in items) if items else False

            if all_exempt:
                # Goes to 3.1(c) - Exempt/Nil rated
                exempt_nil['taxable_value'] += inv.subtotal or 0
            elif inv.igst_total and inv.igst_total > 0 and inv.cgst_total == 0:
                # Inter-state supply - could be export (zero rated) or regular
                # For simplicity, treat inter-state as regular taxable
                taxable_outward['taxable_value'] += inv.subtotal or 0
                taxable_outward['igst'] += inv.igst_total or 0
            else:
                # Regular intra-state taxable supply - 3.1(a)
                taxable_outward['taxable_value'] += inv.subtotal or 0
                taxable_outward['cgst'] += inv.cgst_total or 0
                taxable_outward['sgst'] += inv.sgst_total or 0

    # Deduct credit notes from outward supplies
    cn_total = sum(cn.grand_total or 0 for cn in credit_notes)
    cn_cgst = sum(cn.cgst_total or 0 for cn in credit_notes)
    cn_sgst = sum(cn.sgst_total or 0 for cn in credit_notes)
    cn_igst = sum(cn.igst_total or 0 for cn in credit_notes)

    # 3.2 Inter-state supplies made to
    # (a) Unregistered persons (B2C)
    # (b) Composition taxable persons
    # (c) UIN holders
    b2c_interstate = {'taxable_value': 0}
    for inv in invoices:
        supply_type = getattr(inv, 'supply_type', 'B2C')
        if supply_type == 'B2C' and inv.igst_total and inv.igst_total > 0:
            b2c_interstate['taxable_value'] += inv.subtotal or 0

    # Calculate totals
    total_liability = {
        'igst': taxable_outward['igst'] + reverse_charge['igst'] - cn_igst,
        'cgst': taxable_outward['cgst'] + reverse_charge['cgst'] - cn_cgst,
        'sgst': taxable_outward['sgst'] + reverse_charge['sgst'] - cn_sgst,
        'cess': 0
    }

    return render_template(
        'reports/gstr3b.html',
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
        taxable_outward=taxable_outward,
        zero_rated=zero_rated,
        exempt_nil=exempt_nil,
        reverse_charge=reverse_charge,
        b2c_interstate=b2c_interstate,
        cn_total=cn_total,
        cn_cgst=cn_cgst,
        cn_sgst=cn_sgst,
        cn_igst=cn_igst,
        total_liability=total_liability,
        invoice_count=len(invoices),
        cn_count=len(credit_notes)
    )


@reports_bp.route('/gstr3b/export')
@login_required
def export_gstr3b():
    """Export GSTR-3B to Excel"""
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)

    # Calculate date range
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # Get invoices and credit notes
    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).all()

    credit_notes = CreditNote.query.filter(
        CreditNote.credit_note_date.between(start_date, end_date),
        CreditNote.status != 'CANCELLED'
    ).all()

    # Calculate GSTR-3B data (same logic as above)
    taxable_outward = {'taxable_value': 0, 'igst': 0, 'cgst': 0, 'sgst': 0, 'cess': 0}
    exempt_nil = {'taxable_value': 0}
    reverse_charge = {'taxable_value': 0, 'igst': 0, 'cgst': 0, 'sgst': 0, 'cess': 0}

    for inv in invoices:
        is_rcm = getattr(inv, 'is_reverse_charge', False)
        if is_rcm:
            reverse_charge['taxable_value'] += inv.subtotal or 0
            reverse_charge['igst'] += inv.igst_total or 0
            reverse_charge['cgst'] += inv.cgst_total or 0
            reverse_charge['sgst'] += inv.sgst_total or 0
        else:
            items = list(inv.items) if hasattr(inv, 'items') else []
            all_exempt = all(item.gst_rate == 0 for item in items) if items else False
            if all_exempt:
                exempt_nil['taxable_value'] += inv.subtotal or 0
            else:
                taxable_outward['taxable_value'] += inv.subtotal or 0
                taxable_outward['igst'] += inv.igst_total or 0
                taxable_outward['cgst'] += inv.cgst_total or 0
                taxable_outward['sgst'] += inv.sgst_total or 0

    cn_total = sum(cn.grand_total or 0 for cn in credit_notes)
    cn_cgst = sum(cn.cgst_total or 0 for cn in credit_notes)
    cn_sgst = sum(cn.sgst_total or 0 for cn in credit_notes)
    cn_igst = sum(cn.igst_total or 0 for cn in credit_notes)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-3B Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")

    # Title
    ws.append([f"GSTR-3B Summary - {start_date.strftime('%B %Y')}"])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    # 3.1 Outward Supplies
    ws.append(['3.1 Details of Outward Supplies and Inward Supplies liable to reverse charge'])
    ws['A3'].font = Font(bold=True)
    ws.append([])

    headers = ['Nature of Supplies', 'Taxable Value', 'IGST', 'CGST', 'SGST/UTGST', 'Cess']
    ws.append(headers)
    for col, cell in enumerate(ws[5], 1):
        cell.font = header_font
        cell.fill = header_fill

    ws.append(['(a) Outward taxable supplies', taxable_outward['taxable_value'],
               taxable_outward['igst'], taxable_outward['cgst'], taxable_outward['sgst'], 0])
    ws.append(['(b) Outward taxable supplies (zero rated)', 0, 0, 0, 0, 0])
    ws.append(['(c) Other outward supplies (exempt, nil rated)', exempt_nil['taxable_value'], 0, 0, 0, 0])
    ws.append(['(d) Inward supplies (reverse charge)', reverse_charge['taxable_value'],
               reverse_charge['igst'], reverse_charge['cgst'], reverse_charge['sgst'], 0])
    ws.append([])

    # Credit Notes
    ws.append(['Credit Notes Issued'])
    ws['A11'].font = Font(bold=True)
    ws.append(['Total Credit Notes', cn_total, cn_igst, cn_cgst, cn_sgst, 0])
    ws.append([])

    # Net Liability
    ws.append(['Net Tax Liability'])
    ws['A14'].font = Font(bold=True)
    net_igst = taxable_outward['igst'] + reverse_charge['igst'] - cn_igst
    net_cgst = taxable_outward['cgst'] + reverse_charge['cgst'] - cn_cgst
    net_sgst = taxable_outward['sgst'] + reverse_charge['sgst'] - cn_sgst
    ws.append(['Total Liability', '', net_igst, net_cgst, net_sgst, 0])

    # Format columns
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['A'].width = 45

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month_name = start_date.strftime('%B')
    filename = f"GSTR3B_{month_name}_{year}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/sales/export')
@login_required
def export_sales():
    """Export sales report to Excel"""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).order_by(Invoice.invoice_date, Invoice.invoice_number).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Headers
    headers = [
        "Invoice #", "Date", "Customer", "GSTIN", "Subtotal",
        "CGST", "SGST", "IGST", "Discount", "Grand Total", "Status"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill

    # Data
    for inv in invoices:
        ws.append([
            inv.invoice_number,
            inv.invoice_date.strftime('%d-%m-%Y'),
            inv.customer_name,
            inv.customer.gstin if inv.customer else '',
            inv.subtotal,
            inv.cgst_total,
            inv.sgst_total,
            inv.igst_total,
            inv.discount,
            inv.grand_total,
            'CANCELLED' if inv.is_cancelled else inv.payment_status
        ])

    # Totals row
    total_row = len(invoices) + 2
    ws.append([
        'TOTAL', '', '', '',
        sum(inv.subtotal for inv in invoices),
        sum(inv.cgst_total for inv in invoices),
        sum(inv.sgst_total for inv in invoices),
        sum(inv.igst_total for inv in invoices),
        sum(inv.discount for inv in invoices),
        sum(inv.grand_total for inv in invoices),
        ''
    ])
    for cell in ws[total_row]:
        cell.font = header_font

    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Sales_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/stock/export')
@login_required
def export_stock():
    """Export stock report to Excel"""
    products = Product.query.filter_by(is_active=True).order_by(Product.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    headers = [
        "Product Name", "Barcode", "HSN Code", "Category",
        "Stock", "Unit", "Low Stock Alert", "Price", "Stock Value"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill

    for p in products:
        ws.append([
            p.name,
            p.barcode or '',
            p.hsn_code or '',
            p.category.name if p.category else '',
            p.stock_qty,
            p.unit,
            p.low_stock_alert,
            p.price,
            p.stock_qty * p.price
        ])

    # Totals
    total_row = len(products) + 2
    ws.append([
        'TOTAL', '', '', '', '', '', '', '',
        sum(p.stock_qty * p.price for p in products)
    ])
    for cell in ws[total_row]:
        cell.font = header_font

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Stock_Report_{date.today().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/aging')
@login_required
def customer_aging_report():
    """Customer aging report - outstanding payments by age"""
    # Get all unpaid/partial invoices
    invoices = Invoice.query.filter(
        Invoice.is_cancelled == False,
        Invoice.payment_status.in_(['UNPAID', 'PARTIAL']),
        Invoice.balance_due > 0
    ).order_by(Invoice.invoice_date).all()

    today = date.today()

    # Group by aging buckets
    aging_data = {
        'current': [],      # 0-30 days
        'days_31_60': [],   # 31-60 days
        'days_61_90': [],   # 61-90 days
        'over_90': []       # >90 days
    }

    customer_totals = {}

    for inv in invoices:
        days_old = (today - inv.invoice_date).days if inv.invoice_date else 0

        # Determine bucket
        if days_old <= 30:
            bucket = 'current'
        elif days_old <= 60:
            bucket = 'days_31_60'
        elif days_old <= 90:
            bucket = 'days_61_90'
        else:
            bucket = 'over_90'

        inv_data = {
            'invoice': inv,
            'days_old': days_old,
            'bucket': bucket
        }
        aging_data[bucket].append(inv_data)

        # Track customer totals
        customer_name = inv.customer_name or 'Walk-in'
        if customer_name not in customer_totals:
            customer_totals[customer_name] = {
                'name': customer_name,
                'current': 0,
                'days_31_60': 0,
                'days_61_90': 0,
                'over_90': 0,
                'total': 0
            }
        customer_totals[customer_name][bucket] += inv.balance_due
        customer_totals[customer_name]['total'] += inv.balance_due

    # Summary
    totals = {
        'current': sum(inv.balance_due for inv in [d['invoice'] for d in aging_data['current']]),
        'days_31_60': sum(inv.balance_due for inv in [d['invoice'] for d in aging_data['days_31_60']]),
        'days_61_90': sum(inv.balance_due for inv in [d['invoice'] for d in aging_data['days_61_90']]),
        'over_90': sum(inv.balance_due for inv in [d['invoice'] for d in aging_data['over_90']])
    }
    totals['total'] = sum(totals.values())

    # Sort customers by total outstanding
    customer_list = sorted(customer_totals.values(), key=lambda x: x['total'], reverse=True)

    return render_template(
        'reports/aging.html',
        aging_data=aging_data,
        customer_totals=customer_list,
        totals=totals,
        invoice_count=len(invoices)
    )


@reports_bp.route('/top-products')
@login_required
def top_products_report():
    """Top selling products report"""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    limit = request.args.get('limit', 20, type=int)

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Top products by revenue
    top_by_revenue = db.session.query(
        InvoiceItem.product_name,
        InvoiceItem.hsn_code,
        func.sum(InvoiceItem.qty).label('total_qty'),
        func.sum(InvoiceItem.taxable_value).label('total_taxable'),
        func.sum(InvoiceItem.total).label('total_revenue'),
        func.count(InvoiceItem.id).label('times_sold')
    ).join(Invoice).filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).group_by(InvoiceItem.product_name, InvoiceItem.hsn_code)\
     .order_by(func.sum(InvoiceItem.total).desc()).limit(limit).all()

    # Top products by quantity
    top_by_qty = db.session.query(
        InvoiceItem.product_name,
        InvoiceItem.hsn_code,
        func.sum(InvoiceItem.qty).label('total_qty'),
        func.sum(InvoiceItem.total).label('total_revenue'),
        func.count(InvoiceItem.id).label('times_sold')
    ).join(Invoice).filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).group_by(InvoiceItem.product_name, InvoiceItem.hsn_code)\
     .order_by(func.sum(InvoiceItem.qty).desc()).limit(limit).all()

    # Summary
    total_revenue = sum(p.total_revenue for p in top_by_revenue)
    total_qty = sum(p.total_qty for p in top_by_qty)

    return render_template(
        'reports/top_products.html',
        top_by_revenue=top_by_revenue,
        top_by_qty=top_by_qty,
        start_date=start_date,
        end_date=end_date,
        total_revenue=total_revenue,
        total_qty=total_qty,
        limit=limit
    )


@reports_bp.route('/customer-sales')
@login_required
def customer_sales_report():
    """Customer-wise sales analysis"""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Customer sales summary
    customer_sales = db.session.query(
        Invoice.customer_id,
        Invoice.customer_name,
        func.count(Invoice.id).label('invoice_count'),
        func.sum(Invoice.subtotal).label('total_subtotal'),
        func.sum(Invoice.cgst_total + Invoice.sgst_total + Invoice.igst_total).label('total_tax'),
        func.sum(Invoice.grand_total).label('total_sales'),
        func.avg(Invoice.grand_total).label('avg_invoice')
    ).filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).group_by(Invoice.customer_id, Invoice.customer_name)\
     .order_by(func.sum(Invoice.grand_total).desc()).all()

    # Get customer details
    customer_data = []
    for cs in customer_sales:
        customer = Customer.query.get(cs.customer_id) if cs.customer_id else None
        customer_data.append({
            'customer': customer,
            'name': cs.customer_name or 'Walk-in',
            'invoice_count': cs.invoice_count,
            'total_subtotal': cs.total_subtotal or 0,
            'total_tax': cs.total_tax or 0,
            'total_sales': cs.total_sales or 0,
            'avg_invoice': cs.avg_invoice or 0
        })

    # Summary
    total_sales = sum(c['total_sales'] for c in customer_data)
    total_invoices = sum(c['invoice_count'] for c in customer_data)
    unique_customers = len(customer_data)

    return render_template(
        'reports/customer_sales.html',
        customer_data=customer_data,
        start_date=start_date,
        end_date=end_date,
        total_sales=total_sales,
        total_invoices=total_invoices,
        unique_customers=unique_customers
    )


@reports_bp.route('/profit-loss')
@login_required
def profit_loss_report():
    """Simple Profit & Loss report"""
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    today = date.today()
    if not start_date_str:
        # Default to current financial year
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
        else:
            start_date = date(today.year - 1, 4, 1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Revenue from invoices
    invoices = Invoice.query.filter(
        Invoice.invoice_date.between(start_date, end_date),
        Invoice.is_cancelled == False
    ).all()

    gross_sales = sum(inv.subtotal for inv in invoices)
    total_tax_collected = sum(inv.cgst_total + inv.sgst_total + inv.igst_total for inv in invoices)
    total_discount = sum(inv.discount for inv in invoices)
    net_sales = sum(inv.grand_total for inv in invoices)

    # Credit notes (returns)
    credit_notes = CreditNote.query.filter(
        CreditNote.credit_note_date.between(start_date, end_date),
        CreditNote.status != 'CANCELLED'
    ).all()

    returns_value = sum(cn.grand_total for cn in credit_notes)

    # Cost of goods sold (based on purchase price if available)
    # For now, we'll estimate based on products
    cogs = 0
    for inv in invoices:
        for item in inv.items:
            if item.product_id:
                product = Product.query.get(item.product_id)
                if product and hasattr(product, 'cost_price') and product.cost_price:
                    cogs += product.cost_price * item.qty

    # Gross profit
    gross_profit = net_sales - returns_value - cogs

    # Operating metrics
    invoice_count = len(invoices)
    avg_invoice_value = net_sales / invoice_count if invoice_count > 0 else 0

    return render_template(
        'reports/profit_loss.html',
        start_date=start_date,
        end_date=end_date,
        gross_sales=gross_sales,
        total_discount=total_discount,
        net_sales=net_sales,
        returns_value=returns_value,
        cogs=cogs,
        gross_profit=gross_profit,
        total_tax_collected=total_tax_collected,
        invoice_count=invoice_count,
        avg_invoice_value=avg_invoice_value
    )
